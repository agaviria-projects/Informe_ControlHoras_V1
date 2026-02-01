import tkinter as tk
from tkinter import messagebox
from pathlib import Path
import pandas as pd
import sqlite3
from app.data.db import get_connection

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.data.repository import buscar_registros


def exportar_a_excel(parent):
    try:
        # ======================================================
        # 1) OBTENER TODOS LOS REGISTROS
        # ======================================================
        conn = get_connection()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        cur.execute("""
            SELECT
                cedula,
                nombre,
                placa,
                COALESCE(zona,'') AS zona,
                fecha,
                kilometro,
                horas_trabajadas,
                valor_hora_extra
            FROM registros
            WHERE deleted = 0
        """)
        rows = cur.fetchall()
        conn.close()


        if not rows:
            messagebox.showwarning(
                "Atención",
                "No hay registros para exportar.",
                parent=parent
            )
            return

        # ======================================================
        # 2) DATOS SOLO PARA EL USUARIO (SIN id / created_at)
        # ======================================================
        data = []
        for r in rows:
            data.append([
                r["cedula"],
                r["nombre"],
                r["placa"],
                r["zona"],
                r["fecha"],
                r["kilometro"],
                r["horas_trabajadas"],
                r["valor_hora_extra"],
            ])

        columnas = [
            "Cédula",
            "Nombre",
            "Placa",
            "Zona",
            "Fecha",
            "Kilómetros",
            "Horas trabajadas",
            "Valor Hora Extra",
        ]

        df = pd.DataFrame(data, columns=columnas)

        # ======================================================
        # 3) RUTA /reports
        # ======================================================
        base_dir = Path(__file__).resolve().parents[2]
        reports_dir = base_dir / "reports"
        reports_dir.mkdir(exist_ok=True)

        out = reports_dir / "reporte_control_horas.xlsx"

        # ======================================================
        # 4) EXPORTAR DATAFRAME
        # ======================================================
        df.to_excel(out, index=False)

        # ======================================================
        # 5) CONVERTIR A TABLA ESTRUCTURADA
        # ======================================================
        wb = load_workbook(out)
        ws = wb.active

        last_row = ws.max_row
        last_col = ws.max_column

        col_letter = chr(64 + last_col)
        table_ref = f"A1:{col_letter}{last_row}"

        tabla = Table(
            displayName="TablaControlHoras",
            ref=table_ref
        )

        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)

        # ======================================================
        # 5.1) AUTOAJUSTAR ANCHO DE COLUMNAS (como Excel "AutoFit")
        # ======================================================
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter  # A, B, C...

            for cell in col_cells:
                value = cell.value
                if value is None:
                    continue
                value_len = len(str(value))
                if value_len > max_len:
                    max_len = value_len

            # margen + límite para que no quede gigante
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        wb.save(out)

        # ======================================================
        # 6) MENSAJE FINAL
        # ======================================================
        messagebox.showinfo(
            "Éxito",
            f"Excel generado como tabla estructurada:\n{out}",
            parent=parent
        )
        return out 
    except Exception as e:
        messagebox.showerror(
            "Error",
            f"Ocurrió un error:\n{e}",
            parent=parent
        )
        return None

